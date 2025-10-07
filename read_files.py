import os
import re
import glob
import PyPDF2
import arabic_reshaper
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# ✅ إصلاح الكتابة العربية
def fix_arabic(text: str) -> str:
    try:
        return arabic_reshaper.reshape(text)
    except Exception:
        return text


# ✅ قراءة محتوى كل PDF كسطور نصية
def read_pdf(path: str) -> list:
    context = []
    with open(path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text = page.extract_text()
            if text:
                context.extend(text.splitlines())
    return context


# ✅ استخراج البيانات المطلوبة من النص
def filters(text: list[str]) -> dict:
    due_date, end_of_payments, amount = [], [], []
    row = {}

    for i in range(len(text)):
        if "Contract No" in text[i]:
            row["Contract No"] = text[i].split("Contract No")[1].split(":")[0].replace(". ", "")

        if "Tenancy Start Date" in text[i]:
            row["Tenancy Start Date"] = text[i].split("Tenancy Start Date")[1].split(":")[0]

        if "Tenancy End Date" in text[i]:
            row["Tenancy End Date"] = text[i].split("Tenancy End Date")[1].split(":")[0]

        if "name/Founder" in text[i]:
            company_name = "".join(text[i:i + 2]).split("Organization")[0]
            company_name = company_name.rsplit(" ", 1)[0].replace("name/Founder", "")[:-3]
            row["Tenancy Name"] = fix_arabic(company_name)

        if "National Address" in text[i]:
            row["National Address"] = text[i].replace("National Address", "")

        if "Lessor Data" in text[i]:
            try:
                lessor_text = " ".join(text[i:i + 3])
                lessor_text = re.sub(r"\s+", " ", lessor_text).strip()
                if "Name" in lessor_text:
                    name_part = lessor_text.split("Name", 1)[1].strip()
                    name_clean = re.sub(r"^[:\s]*الاسم[:\s]*", "", name_part).strip()
                    name_clean = re.split(r"Nationality|:", name_clean)[0].strip()
                    name_clean = re.sub(r"(اﻻﺳﻢ|الاسم|\sالاسم|\sاﻻﺳﻢ)$", "", name_clean, flags=re.UNICODE).strip()
                    row["Lessor Name"] = fix_arabic(name_clean)
            except Exception:
                row["Lessor Name"] = ""

        # استخراج جدول الدفعات
        pattern = r"^\d+\.\d+\s+\d{4}-\d{2}-\d{2}\s+\d{4}-\d{2}-\d{2}.*\d{4}-\d{2}-\d{2}\s+\d{4}-\d{2}-\d{2}\s+\d+\s*$"
        payments = re.findall(pattern, text[i])
        if payments:
            payments = "".join(payments).split(" ")
            due_date.append(payments[5])
            end_of_payments.append(payments[4])
            amount.append(payments[0])

    row["Due Date"] = due_date[:]
    row["End of Payments"] = end_of_payments[:]
    row["Amount"] = amount[:]
    return row


# ✅ إضافة البيانات إلى ملف Excel (أو إنشاءه إن لم يوجد)
def convert_to_excel(data, output_file: str) -> None:
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contract Data"
        headers = [
            "Contract No",
            "Tenancy Start Date",
            "Tenancy End Date",
            "Tenancy Name",
            "National Address",
            "Lessor Name",
            "Due Date",
            "End of Payments",
            "Amount"
        ]
        ws.append(headers)

    start_row = ws.max_row + 1

    ws.append([
        data.get("Contract No", ""),
        data.get("Tenancy Start Date", ""),
        data.get("Tenancy End Date", ""),
        data.get("Tenancy Name", ""),
        data.get("National Address", ""),
        data.get("Lessor Name", ""),
        "", "", ""
    ])

    row_index = start_row + 1
    for due, end, amount in zip(data.get('Due Date', []),
                                data.get('End of Payments', []),
                                data.get('Amount', [])):
        ws.cell(row=row_index, column=7, value=due)
        ws.cell(row=row_index, column=8, value=end)
        ws.cell(row=row_index, column=9, value=amount)
        row_index += 1
    ws.append([""] * 9)
    ws.append([""] * 9)
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                cell_length = len(str(cell.value))
                if col_letter not in column_widths:
                    column_widths[col_letter] = cell_length
                else:
                    column_widths[col_letter] = max(column_widths[col_letter], cell_length)
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width + 2

    wb.save(output_file)
    print(f"✅ Added to Excel: {output_file}")



pdf_folder = r"C:\Users\ream8\Desktop\project\PDFReaderProject"
excel_path = os.path.join(pdf_folder, "Tenant_Info.xlsx")

pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
print(f"Found {len(pdf_files)} PDF files.")

for pdf_path in pdf_files:
    print(f"Processing: {os.path.basename(pdf_path)}")
    extracting = read_pdf(pdf_path)
    pdf_data = filters(extracting)
    convert_to_excel(pdf_data, excel_path)

print("All PDFs processed successfully!")
